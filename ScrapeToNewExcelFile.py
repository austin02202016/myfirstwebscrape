from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import openpyxl
import re


def scrape_epa_data(link):

    driver = webdriver.Chrome()
    try:
        driver.get(link)
        page_title = driver.title
        
        # Find the specific table by its ID
        table = driver.find_element(By.ID, "listing")
        print(table)
        
        # Extract the table headers
        headers = [header.text.strip() for header in table.find_elements(By.TAG_NAME, "th")]
        print(headers)
        
        # Extract the table rows
        rows = table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip the header row
        
        # Initialize lists to hold the data
        data_rows = []
        
        # Loop through each row in the table
        for row in rows:
            cols = row.find_elements(By.TAG_NAME, "td")
            cols_text = [col.text.strip() for col in cols]  # Extract and strip text from each cell
            
            # Append the row data to the data_rows list
            data_rows.append(cols_text)
        
        # Create a DataFrame from the extracted data
        df = pd.DataFrame(data_rows, columns=headers)
        
        # Export the DataFrame to an Excel file
        df.to_excel("leads_list.xlsx", index=False)
        size = len(data_rows)
        print(size)
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        driver.quit()

def append_data_to_excel(link, filename):
    driver = webdriver.Chrome()
    try:
        driver.get(link)
        page_title = driver.title

        # Find the specific table by its ID
        table = driver.find_element(By.ID, "listing")

        # Extract the table headers
        headers = [header.text.strip() for header in table.find_elements(By.TAG_NAME, "th")]

        # Extract the table rows
        rows = table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip the header row

        # Initialize lists to hold the data
        data_rows = []

        # Loop through each row in the table
        for row in rows:
            cols = row.find_elements(By.TAG_NAME, "td")
            cols_text = [col.text.strip() for col in cols]  # Extract and strip text from each cell

            # Append the row data to the data_rows list
            data_rows.append(cols_text)

        # Create a DataFrame from the extracted data
        df = pd.DataFrame(data_rows, columns=headers)

        # Open the existing Excel file and append the new data
        with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # Check if the sheet exists and append; otherwise, write a new sheet
            if page_title in writer.book.sheetnames:
                startrow = writer.book[page_title].max_row
                df.to_excel(writer, sheet_name=page_title, startrow=startrow, index=False, header=False)
            else:
                df.to_excel(writer, sheet_name=page_title, index=False)
                
        print(f"Appended data to {filename}")

    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        driver.quit()

def re100_scraper(link, filename):

    driver = webdriver.Chrome()
    try:
        driver.get(link)
        page_title = driver.title

        
        # Find the specific table by its ID
        table = driver.find_element(By.CLASS_NAME, "cols-9")
        print(table)
        
        # Extract the table headers
        headers = [header.text.strip() for header in table.find_elements(By.TAG_NAME, "th")]
        print(headers)
        
        # Extract the table rows
        rows = table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip the header row
        
        # Initialize lists to hold the data
        data_rows = []
        
        # Loop through each row in the table
        for row in rows:
            cols = row.find_elements(By.TAG_NAME, "td")
            cols_text = [col.text.strip() for col in cols]  # Extract and strip text from each cell
            
            # Append the row data to the data_rows list
            data_rows.append(cols_text)
        
        # Create a DataFrame from the extracted data
        df = pd.DataFrame(data_rows, columns=headers)
        
        # Open the existing Excel file and append the new data
        with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # Check if the sheet exists and append; otherwise, write a new sheet
            if page_title in writer.book.sheetnames:
                startrow = writer.book[page_title].max_row
                df.to_excel(writer, sheet_name=page_title, startrow=startrow, index=False, header=False)
            else:
                df.to_excel(writer, sheet_name=page_title, index=False)
                
        print(f"Appended data to {filename}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        driver.quit()


def delete_odd_rows(filename, sheet_name):
    # Load the workbook and the specific sheet
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]

    # Delete rows in reverse order to avoid shifting indices
    max_row = 881
    min_row = 3
    for row in range(max_row, min_row - 1, -1):
        if row % 2 != 0:  # Check if the row number is odd
            sheet.delete_rows(row, 1)  # Delete the row

    # Save the modified workbook
    workbook.save(filename)
    print(f"Odd-numbered rows from {min_row} to {max_row} have been deleted from '{sheet_name}'.")

def all_entities():
    # Load the workbook
    workbook = openpyxl.load_workbook('entities_list.xlsx')

    # A set to store unique entities
    entities = set()

    # Iterate through every sheet in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0] is not None:  # Ensure the cell is not empty
                entities.add(row[0])  # Add the value from column A to the set

    # Create a new sheet titled "All Entities"
    if "All Entities" in workbook.sheetnames:
        all_entities_sheet = workbook["All Entities"]  # Use existing sheet
    else:
        all_entities_sheet = workbook.create_sheet("All Entities")  # Create a new sheet

    # Add each unique entity to the new sheet in column A
    for i, entity in enumerate(entities, start=1):
        all_entities_sheet[f'A{i}'] = entity

    # Save the modified workbook
    workbook.save('entities_list.xlsx')
    print("All unique entities have been added to the 'All Entities' sheet.")
